import openpyxl
import pandas as pd

file_path = "DATA IN COMPRAS.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print("--- WORKBOOK ANALYSIS ---")
    
    for sheet_name in wb.sheetnames:
        print(f"\nSheet: {sheet_name}")
        ws = wb[sheet_name]
        
        # Get headers
        headers = [str(cell.value) if cell.value else f"Col_{i}" for i, cell in enumerate(ws[1])]
        print(f"Headers: {headers}")
        
        # Analyze first 10 rows to see data patterns
        rows = list(ws.iter_rows(min_row=2, max_row=11, values_only=True))
        for i, row in enumerate(rows):
            # Filter out completely empty rows
            if any(row):
                print(f"Row {i+1}: {row}")

except Exception as e:
    print(f"Error: {e}")
